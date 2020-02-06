# A program to fulfill the project requirements for Hash Table in class MS549: timing the duration
# of 100 random hash table entries, deletions, and removals

# In this project I not only practice hash tables, but also the use of color within my output, using both class color
# variable assignment, and the library 'termcolor'.

# reference: https://docs.python.org/3/library/hashlib.html
# https://pypi.org/project/hashlib/
# https://readthedocs.org/projects/hashlib/
# https://circuitpython.readthedocs.io/projects/hashlib/en/latest/
# https://github.com/adafruit/Adafruit_CircuitPython_hashlib
# https://www.pythoncentral.io/hashing-strings-with-python/
# Reference for uuid to add salt to hash to prevent collision:
# https://stackoverflow.com/questions/9594125/salt-and-hash-a-password-in-python
# https://docs.python.org/2/library/uuid.html

# Other references for hash tables in python:
# https://www.geeksforgeeks.org/python-hash-method/'
# https://www.geeksforgeeks.org/applications-of-hashing/
# side note from reference:  uuid4() creates a random UUID.

# importing 'hashlib' to create a hash table, and 'uuid' to salt the hash and prevent collision
import hashlib, _sha256, uuid, binascii
import pandas as pd
import numpy as np
import random
import xlrd
import hmac
import hashlib, uuid
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import xlrd
import sys


# To get rid of future warnings, specifically from hashClass(), function _getInfo_():
# FutureWarning: elementwise comparison failed; returning scalar instead, but in the future will perform elementwise comparison
# res_values = method(rvalues)
# reference:  https://stackoverflow.com/questions/40659212/futurewarning-elementwise-comparison-failed-returning-scalar-but-in-the-futur
import warnings
with warnings.catch_warnings():
    warnings.simplefilter(action='ignore', category=FutureWarning)

dataEmpty = bytes([])
print("\nEmpty dataset:", dataEmpty)

class Hashing:

    def newNum(self):

        rand100 = random.sample(range(1, 101), 1)

        randomUpdateNum = random.sample(range(1, 255), 5)
        randomUpdateNumBytes = bytes(randomUpdateNum)

        dataRandom = np.append(dataEmpty, [rand100])
        print("\nPlacing 100 random numbers into the dataset:\n", dataRandom)

        dataRandomBytes = bytearray(dataRandom)
        salt = uuid.uuid4().hex

        for item in dataRandomBytes:

            item1 = bytes(item)
            h = hashlib.new('sha256', item1)
            h.update(randomUpdateNumBytes)
            itemHashed = h.hexdigest()
            itemHashed_and_Salted = itemHashed + ":" + salt

            print("\nAdding a hash to the 100 dataset numbers:\n", itemHashed_and_Salted)

            k = itemHashed_and_Salted
            my_hash_table = {k: dataRandom}

            myNewDataframe = pd.DataFrame.from_dict(my_hash_table, orient='index')
            writer = pd.ExcelWriter('MyHashTable.xlsx', engine='openpyxl')
            writer.book = load_workbook('MyHashTable.xlsx')
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            reader = pd.read_excel(r'MyHashTable.xlsx')
            myNewDataframe.to_excel(writer, index=True, header=False, startrow=len(reader) + 1)
            writer.close()
            return

    def newNumPrint(self):

        h = Hashing()

        # The hashes were all showing up as same hash and salt, so created the function and now calling the function 100 times
        # reference: https://stackoverflow.com/questions/4264634/more-pythonic-way-to-run-a-process-x-times
        # the range (99) is for 0 - 99, which would be 100 numbers
        for item in range(99):
            h.newNum()

        print("\n100 values have been added.\n")

    def _getInfo_(self):

        retrieveInput = input("Please enter your assigned private key in order to retrieve personal employee data.\n\n" +
                              "Key: ")

        MyDataframe = pd.read_excel('MyHashTable.xlsx', sheet_name="Sheet1",
                                    keep_default_na=False, na_values=[""])

        # references for dropna:
        # https://stackoverflow.com/questions/53856763/get-row-and-column-in-pandas-for-a-cell-with-a-certain-value
        # https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.dropna.html
        # reference for iloc:
        # https://www.shanelynn.ie/select-pandas-dataframe-rows-and-columns-using-iloc-loc-and-ix/
        # https: // www.edureka.co / community / 43228 / pandas - iloc - printing - whole - row - instead - of - cell - value

        # Using the dropna() method in python to drop away other rows except for the retrieveInput variable (which is
        # the user's entry of the key).
        a = MyDataframe.where(MyDataframe == retrieveInput).dropna(how='all').dropna(axis=1)

        if a is not None:
            # Getting the key index using the dropna() method assigned variable, and assigning this index the
            # variable newIndex
            newIndex = a.index

            # Now that the index has been located using the dropna() method, the iloc() method can be used
            # to slice away all the other data except for the row in that index which was assigned the variable newIndex
            # reference for using pandas (pd) options for making the table look nice when printed:
            # https://towardsdatascience.com/pretty-displaying-tricks-for-columnar-data-in-python-2fe3b3ed9b83
            pd.options.display.max_columns = None
            pd.options.display.width = None
            keyRow = (MyDataframe.iloc[newIndex])
            print('Printing the requested hashed key and assigned random 100 number value:\n', keyRow, "\n")

        else:
            print("Error")
            return

        # Creating user-input and using the pandas drop() method to allow deletion requests
        deleteRequ = input("Do you wish to delete this input permanently from the record?\nPlease type Yes, or No: ")
        if deleteRequ == "Yes":
            doubleChecking = input("Are you sure? Please type Yes, or No:")
            if doubleChecking == "Yes":

                # Creating a variable for data being pulled from the MyEmployeeHashTable.xlsx file and placed into array format
                book = xlrd.open_workbook('MyHashTable.xlsx')
                sheets = book.sheets()
                for sheet in sheets:
                    data = np.array([[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)])

                # Creating a variable for data to be transformed into excel file format for writing on file
                # references utilized:
                # https://www.shanelynn.ie/using-pandas-dataframe-creating-editing-viewing-data-in-python/
                # https://stackoverflow.com/questions/44931834/pandas-drop-function-error-label-not-contained-in-axis
                writer = pd.ExcelWriter(book)
                writer.book = load_workbook('MyHashTable.xlsx')
                writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                reader = pd.read_excel(r'MyHashTable.xlsx', index_col=0)
                reader.drop(retrieveInput, axis=0, inplace=True)
                reader.to_excel('MyHashTable.xlsx')  ##, index=True, header=False, startrow=len(reader)-1)
                writer.close()

                print("The file has been permanently deleted.")

            if doubleChecking == "No":
                print("We will return you to the main menu.")
                return
            else:
                return
        if deleteRequ == "No":
            print("We will return you to the main menu.")
            return

        else:
            return
Hashing()

def main():

    hasher = Hashing()

    asking = input("Please make a choice.\n"+
                   "1 to make an entry of 100 numbers onto the excel file\n" +
                   "2 to retrieve, and if you wish, remove a number from the excel file \n" +
                   "3 to empty the excel file\n"
                   "4 to exit the program\n" +
                   "\n\nMake your entry here: ")

    if asking == '1':
        hasher.newNumPrint()

    if asking == '2':
        hasher._getInfo_()

main()





