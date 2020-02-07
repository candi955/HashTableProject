# A program to fulfill the project requirements for Hash Table in class MS549: timing the duration
# of 10,000 random hash table entries, deletions, and removals
# Initially, a menu shows (menu() function at bottom of program) to offer the user choices

# Of special note: the time duration to complete 10,000 insert iteration of the hash and salt and place them in the
# excel file, for this project, estimated at approximately 1.34 seconds (for 9,378 hash/salt iterations it took
# approximately 219 minutes, or 12,600 seconds to run).

# Due to this, the estimate of time duration for completion of this insert function, in
# this project, would be as follows:
# For 10,000 iterations:
# 13,400 seconds
# 223 minutes
# 3.72 hours

# In contrast, deletion time duration for 'delete all' from the DataFrame is still 0 seconds.



# reference: https://docs.python.org/3/library/hashlib.html
# https://pypi.org/project/hashlib/
# https://readthedocs.org/projects/hashlib/
# https://circuitpython.readthedocs.io/projects/hashlib/en/latest/
# https://github.com/adafruit/Adafruit_CircuitPython_hashlib
# https://www.pythoncentral.io/hashing-strings-with-python/
# Reference for uuid to add salt to hash to prevent collision:
# https://stackoverflow.com/questions/9594125/salt-and-hash-a-password-in-python
# https://docs.python.org/2/library/uuid.html

# Other references for hash tables in python:
# https://www.geeksforgeeks.org/python-hash-method/'
# https://www.geeksforgeeks.org/applications-of-hashing/
# side note from reference:  uuid4() creates a random UUID.

# importing 'hashlib' to create a hash table, and 'uuid' to salt the hash and prevent collision
import hashlib, uuid
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import xlrd
import random
import time

# To get rid of future warnings, specifically from hashClass(), function _getInfo_():
# FutureWarning: elementwise comparison failed; returning scalar instead, but in the future will perform elementwise comparison
# res_values = method(rvalues)
# reference:  https://stackoverflow.com/questions/40659212/futurewarning-elementwise-comparison-failed-returning-scalar-but-in-the-futur
import warnings
with warnings.catch_warnings():
    warnings.simplefilter(action='ignore', category=FutureWarning)

dataEmpty = bytes([])
print("\nEmpty dataset:", dataEmpty)

class Hashing:

    def newNum(self):

        rand10000 = random.sample(range(1, 10001), 1)

        randomUpdateNum = random.sample(range(1, 255), 5)
        randomUpdateNumBytes = bytes(randomUpdateNum)

        dataRandom = np.append(dataEmpty, [rand10000])
        print("\nPlacing 10,000 random numbers into the dataset:\n", dataRandom)

        dataRandomBytes = bytearray(dataRandom)
        salt = uuid.uuid4().hex

        for item in dataRandomBytes:

            item1 = bytes(item)
            h = hashlib.new('sha256', item1)
            h.update(randomUpdateNumBytes)
            itemHashed = h.hexdigest()
            itemHashed_and_Salted = itemHashed + ":" + salt

            print("\nAdding a hash to the 10,000 dataset numbers:\n", itemHashed_and_Salted)

            k = itemHashed_and_Salted
            my_hash_table = {k: dataRandom}

            myNewDataframe = pd.DataFrame.from_dict(my_hash_table, orient='index')
            writer = pd.ExcelWriter('MyHashTable10_000.xlsx', engine='openpyxl')
            writer.book = load_workbook('MyHashTable10_000.xlsx')
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            reader = pd.read_excel(r'MyHashTable10_000.xlsx')
            myNewDataframe.to_excel(writer, index=True, header=False, startrow=len(reader) + 1)
            writer.close()
            return

    def newNumPrint(self):

        h = Hashing()

        # Creating start, stop, and duration variables to time how long the functions take
        timeToHashStart = time.time()

        # The hashes were all showing up as same hash and salt, so created the function and now calling the function 10,000 times
        # reference: https://stackoverflow.com/questions/4264634/more-pythonic-way-to-run-a-process-x-times
        # the range (10000) at the current setup within this python version and IDE, is creating 10,000 hashes
        # within the newNum() function and newNumPrint() function for loops.
        for item in range(10000):
            h.newNum()

        timeToHashEnd = time.time()
        timeToHashDuration = timeToHashEnd - timeToHashStart

        print("\n10,000 values have been added.\n")
        print("\nThe timing it took to add 10,000 random values with hashes/salt to the excel file is",
              timeToHashDuration,"seconds\n")

    def _getInfo_(self):

        retrieveInput = input("Please enter your assigned private key in order to retrieve personal employee data.\n\n" +
                              "Key: ")

        MyDataframe = pd.read_excel('MyHashTable10_000.xlsx', sheet_name="Sheet1",
                                    keep_default_na=False, na_values=[""])

        # references for dropna:
        # https://stackoverflow.com/questions/53856763/get-row-and-column-in-pandas-for-a-cell-with-a-certain-value
        # https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.dropna.html
        # reference for iloc:
        # https://www.shanelynn.ie/select-pandas-dataframe-rows-and-columns-using-iloc-loc-and-ix/
        # https: // www.edureka.co / community / 43228 / pandas - iloc - printing - whole - row - instead - of - cell - value

        # Using the dropna() method in python to drop away other rows except for the retrieveInput variable (which is
        # the user's entry of the key).
        a = MyDataframe.where(MyDataframe == retrieveInput).dropna(how='all').dropna(axis=1)

        # Creating start, stop, and duration variables to time how long the functions take
        timeToRetrieveStart = time.time()

        if a is not None:
            # Getting the key index using the dropna() method assigned variable, and assigning this index the
            # variable newIndex
            newIndex = a.index

            # Now that the index has been located using the dropna() method, the iloc() method can be used
            # to slice away all the other data except for the row in that index which was assigned the variable newIndex
            # reference for using pandas (pd) options for making the table look nice when printed:
            # https://towardsdatascience.com/pretty-displaying-tricks-for-columnar-data-in-python-2fe3b3ed9b83
            pd.options.display.max_columns = None
            pd.options.display.width = None
            keyRow = (MyDataframe.iloc[newIndex])
            print('Printing the requested hashed key and assigned random 10,000 number value:\n', keyRow, "\n")

            timeToRetrieveEnd = time.time()
            timeToRetrieveDuration = timeToRetrieveEnd - timeToRetrieveStart

            print("\nThe timing it took to retrieve a random value with hash/salt from the excel file is",
                  timeToRetrieveDuration, "seconds\n")

        else:
            print("Error")
            return

        # Creating user-input and using the pandas drop() method to allow deletion requests
        deleteRequ = input("\nDo you wish to delete this input permanently from the record?\nPlease type Yes, or No: ")
        if deleteRequ == "Yes":
            doubleChecking = input("\nAre you sure? Please type Yes, or No:")
            if doubleChecking == "Yes":

                # Creating a variable for data being pulled from the MyEmployeeHashTable10_000.xlsx file and placed into array format
                book = xlrd.open_workbook('MyHashTable10_000.xlsx')
                sheets = book.sheets()
                for sheet in sheets:
                    data = np.array([[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)])

                # Creating a variable for data to be transformed into excel file format for writing on file
                # references utilized:
                # https://www.shanelynn.ie/using-pandas-dataframe-creating-editing-viewing-data-in-python/
                # https://stackoverflow.com/questions/44931834/pandas-drop-function-error-label-not-contained-in-axis
                writer = pd.ExcelWriter(book)
                writer.book = load_workbook('MyHashTable10_000.xlsx')
                writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                reader = pd.read_excel(r'MyHashTable10_000.xlsx', index_col=0)
                reader.drop(retrieveInput, axis=0, inplace=True)
                reader.to_excel('MyHashTable10_000.xlsx')  ##, index=True, header=False, startrow=len(reader)-1)
                writer.close()

                print("\nThe file has been permanently deleted.\n")

            if doubleChecking == "No":
                print("\nWe will return you to the main menu.\n")
                return
            else:
                return
        if deleteRequ == "No":
            print("We will return you to the main menu.")
            return

        else:
            return

    def _deletingAll_(self):

        MyDataframe = pd.read_excel('MyHashTable10_000.xlsx', sheet_name="Sheet1",
                                    keep_default_na=False, na_values=[""])

        # Creating user-input and using the pandas drop() method to allow deletion requests
        deleteRequ = input("Do you wish to delete this input permanently from this entire dataframe?" +
                           "\nPlease type Yes, or No: ")
        if deleteRequ == "Yes":
            doubleChecking = input("Are you sure? Please type Yes, or No:")
            if doubleChecking == "Yes":

                # Creating a variable for data being pulled from the MyEmployeeHashTable10_000.xlsx file and placed into array format
                book = xlrd.open_workbook('MyHashTable10_000.xlsx')
                sheets = book.sheets()
                for sheet in sheets:
                    data = np.array([[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)])

                # Creating a variable for data to be transformed into excel file format for writing on file
                # references utilized:
                # https://www.shanelynn.ie/using-pandas-dataframe-creating-editing-viewing-data-in-python/
                # https://stackoverflow.com/questions/44931834/pandas-drop-function-error-label-not-contained-in-axis
                writer = pd.ExcelWriter(book)
                writer.book = load_workbook('MyHashTable10_000.xlsx')
                writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                reader = pd.read_excel(r'MyHashTable10_000.xlsx', index_col=0)
                if reader.empty == False:

                    # Creating start, stop, and duration variables to time how long the functions take
                    timeToDeleteStart = time.time()

                    reader = pd.DataFrame([])

                    timeToDeleteEnd = time.time()
                    timeToDeleteDuration = timeToDeleteEnd - timeToDeleteStart

                    reader.to_excel('MyHashTable10_000.xlsx')
                    writer.close()

                    print("\nThe file has been permanently deleted.\n")

                    print("\nThe timing it took to delete 10,000 random values with hashes/salt from the excel file is",
                          timeToDeleteDuration, "seconds\n")

                    main()
                else:
                    print("\nThis file was empty. Nothing was deleted.\n")
                    main()

            if doubleChecking == "No":
                print("\nWe will return you to the main menu.\n")
                return
            else:
                return
        if deleteRequ == "No":
            print("We will return you to the main menu.")
            return

        else:
            return

Hashing()

def main():

    hasher = Hashing()

    asking = input("Please make a choice.\n"+
                   "1 to make an entry of 10,000 numbers onto the excel file\n" +
                   "2 to retrieve, and if you wish, remove a number from the excel file \n" +
                   "3 to empty the excel file\n"
                   "4 to exit the program\n" +
                   "\n\nMake your entry here: ")

    if asking == '1':
        hasher.newNumPrint()
    if asking == '2':
        hasher._getInfo_()
    if asking == '3':
        hasher._deletingAll_()
    if asking == '4':
        print("\n----Exiting the program----\n")
    else:
        main()

main()





