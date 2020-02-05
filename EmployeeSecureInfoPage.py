# This page is meant to contribute to my final project for the course MS549 Data Structures and Testing
# The program is an example of secure employee data entry using a Hash Table
# The user enters the secure data upon program input prompts, the data is hashed (SHA-256 algorithm) with a salt added,
# utilizing the python library hashlib, and then appended to an excel file.
# The adding of salt contributes to the prevention of collision.

# importing 'hashlib' to create a hash table, and 'uuid' to salt the hash and prevent collision
# Other reference for possible future ideas:
# For checking integrity of the data and user-entries: https://www.pythoncentral.io/hashing-strings-with-python/
# Hash compare online reference: http://onlinemd5.com/
import hashlib, uuid
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import xlrd

# To get rid of future warnings, specifically from hashClass(), function _getInfo_():
# FutureWarning: elementwise comparison failed; returning scalar instead, but in the future will perform elementwise comparison
# res_values = method(rvalues)
# reference:  https://stackoverflow.com/questions/40659212/futurewarning-elementwise-comparison-failed-returning-scalar-but-in-the-futur
import warnings
with warnings.catch_warnings():
    warnings.simplefilter(action='ignore', category=FutureWarning)

class hashClass():

    # Creating the hashClass() constructor
    # Received the idea from the following reference:
    # https://www.journaldev.com/17357/python-hash-function
    def _init_(self, lastName, firstName, empNum, newDOB):
        self.lastName = lastName
        self.firstName = firstName
        self.empNum = empNum
        self.newDOB = newDOB

    # Creating a function for the program user-inputs of employee information, and then to create the table
    # and individual hash keys, added together for an ultimate individual key.
    def _updatedInfo_(self):

        # Creating variables for user-input of employee information into the system
        self.lastName = input("\nPlease enter your last name: ")
        self.firstName = input("\nPlease enter your first name: ")
        self.empNum = input("\nPlease enter your employee number: ")
        self.newDOB = input("\nPlease enter your date-of-birth(DDMMMYYYY: ")

        # Showing the user their input entries for the hash table values
        print("Please check your entries for accuracy, and take personal note of your secure key:\n",
              self.lastName, ", ", self.firstName, ", ", self.empNum, ", ", self.newDOB, "\n")

        # Saving other hash:salt variables for employee information in case needed later to expand program
        ### w = self.lastName
        ### x = self.firstName
        y = self.empNum
        ### z = self.newDOB

        # reference for creating hashes with salt for all of the individual user-inputs
        # https: // python.readthedocs.io / en / latest / library / hashlib.html
        # reference for fixing error concerning hashing a variable (w, x, y, z):
        # https://stackoverflow.com/questions/24905062/how-to-hash-a-variable-in-python

        # reference for adding salt to the hash: https://gist.github.com/markito/30a9bc2afbbfd684b31986c2de305d20
        salt = uuid.uuid4().hex

        # variables for hashing the employee number and adding salt (which will be used in the initially set-up program
        dkEmpNum = hashlib.sha256(y.encode('utf-8'))
        dkEmpNumHashed = dkEmpNum.hexdigest() + ':' + salt
        # Saved other variables for hashing employee information in case wished to use them within the program
        # on a later date
        # variables for hashing the Last Name and adding salt
        ###LastNameString = hashlib.sha256(w.encode('utf-8'))
        ###LastNameHashed = LastNameString.hexdigest() + ':' + salt
        # variables for hashing the First Name and adding salt
        ###dkFirstName = hashlib.sha256(w.encode('utf-8'))
        ###dkFirstNameHashed=dkFirstName.hexdigest() + ':' + salt
        # variables for hashing the employee DOB and adding salt
        ###dkDOB = hashlib.sha256(w.encode('utf-8'))
        ###dkDOBHashed = dkDOB.hexdigest() + ':' + salt

        # Creating a variable called k (for key), to add all of the hashes together to create an
        # individual hashtable key and reference for all of the actual data
        k = dkEmpNumHashed ###LastNameHashed + dkFirstNameHashed + dkEmpNumHashed + dkDOBHashed

        # reference for tuple information:
        # https://www.quora.com/In-the-Python-dictionary-can-1-key-hold-more-than-1-value
        my_table = {k: (self.lastName, self.firstName, self.empNum, self.newDOB)}

        print("\n", my_table)

        # ____________________________________________________________________
        # Attempting to create a dataframe to save the info to a pandas dataframe
        my_table_DataFrame = pd.DataFrame.from_dict(my_table, orient='index')

        # setting the dataframe to show the entire dataframe
        pd.set_option('display.max_rows', 500)
        pd.set_option('display.max_columns', 500)
        pd.set_option('display.width', 1000)

        # printing the dataframe
        print(my_table_DataFrame)

        #____________________________________________________
        # Create, write to and save a workbook
        # https://medium.com/better-programming/using-python-pandas-with-excel-d5082102ca27

        writer = pd.ExcelWriter('MyEmployeeHashTable.xlsx', engine='openpyxl')
        writer.book = load_workbook('MyEmployeeHashTable.xlsx')
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        reader = pd.read_excel(r'MyEmployeeHashTable.xlsx')
        my_table_DataFrame.to_excel(writer, index=True, header=False, startrow=len(reader)+1)
        writer.close()

    # Creating a function to retrieve the associated values with the correct key
    def _getInfo_(self):

        # reference for Confidentiality message creation:
        # https://www.mail-signatures.com/articles/email-disclaimer-examples/
        retrieveInput = input("\nThe content of the data being accessed is confidential and intended for the employee\n" +
              " referenced by the  specified key or legally signed employer and associated Human Resources\n" +
              "departments, or otherwise legally entity with legal authority to access only.\n\n"
              
              "It is strictly forbidden to share any part of this secure data with\n" +
              "any third party, without a written consent of the employee. If you retrieve this data\n" +
              " by mistake, please notify the associated company Human Resources department immediately and\n" +
              "follow with by deleting and discarding any personal records or holdings of the data in\n" +
              "a secure manner, so that we can ensure such a mistake does not occur in the future.\n\n" +

              "Please enter your assigned private key in order to retrieve personal employee data.\n\n" +
              "Key: ")

        MyDataframe = pd.read_excel('MyEmployeeHashTable.xlsx', sheet_name="Sheet1",
                                        keep_default_na=False, na_values=[""])

        # references for dropna:
        # https://stackoverflow.com/questions/53856763/get-row-and-column-in-pandas-for-a-cell-with-a-certain-value
        # https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.dropna.html
        # reference for iloc:
        # https://www.shanelynn.ie/select-pandas-dataframe-rows-and-columns-using-iloc-loc-and-ix/
        # https: // www.edureka.co / community / 43228 / pandas - iloc - printing - whole - row - instead - of - cell - value

        # Using the dropna() method in python to drop away other rows except for the retrieveInput variable (which is
        # the user's entry of the key).
        a = MyDataframe.where(MyDataframe == retrieveInput).dropna(how='all').dropna(axis=1)

        if a is not None:
            # Getting the key index using the dropna() method assigned variable, and assigning this index the
            # variable newIndex
            newIndex = a.index

            # Now that the index has been located using the dropna() method, the iloc() method can be used
            # to slice away all the other data except for the row in that index which was assigned the variable newIndex
            # reference for using pandas (pd) options for making the table look nice when printed:
            # https://towardsdatascience.com/pretty-displaying-tricks-for-columnar-data-in-python-2fe3b3ed9b83
            pd.options.display.max_columns = None
            pd.options.display.width = None
            keyRow = (MyDataframe.iloc[newIndex])
            print('Printing the requested employee key and assigned data:\n', keyRow, "\n")

        else:
            print("Error")
            return

        # Creating user-input and using the pandas drop() method to allow deletion requests
        deleteRequ = input("For administrators only, do you wish to delete this input permanently from the\n" +
                           "record? Please type Yes, or No: ")
        if deleteRequ == "Yes":
            doubleChecking = input("Are you sure? Please type Yes, or No:")
            if doubleChecking == "Yes":

                # Creating a variable for data being pulled from the MyEmployeeHashTable.xlsx file and placed into array format
                book = xlrd.open_workbook('MyEmployeeHashTable.xlsx')
                sheets = book.sheets()
                for sheet in sheets:
                    data = np.array([[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)])

                # Creating a variable for data to be transformed into excel file format for writing on file
                # references utilized:
                # https://www.shanelynn.ie/using-pandas-dataframe-creating-editing-viewing-data-in-python/
                # https://stackoverflow.com/questions/44931834/pandas-drop-function-error-label-not-contained-in-axis
                writer = pd.ExcelWriter(book)
                writer.book = load_workbook('MyEmployeeHashTable.xlsx')
                writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                reader = pd.read_excel(r'MyEmployeeHashTable.xlsx', index_col=0)
                reader.drop(retrieveInput, axis=0, inplace=True)
                reader.to_excel('MyEmployeeHashTable.xlsx') ##, index=True, header=False, startrow=len(reader)-1)
                writer.close()

                print("The file has been permanently deleted.")

            if doubleChecking == "No":
                print("We will return you to the main menu.")
                return
            else:
                return
        if deleteRequ == "No":
            print("We will return you to the main menu.")
            return

        else:
            return


hashClass()

def main():
    # Creating a variable to call the hashClass(), and using that to call the hashClass functions
    theHash = hashClass()

    tryAgain = input("\nPlease enter  on of the following choices:\n" +
                     "1 to make an employee data entry into the Employee Secure Info Page\n" +
                     "2 to retrieve secure data using the key\n" +
                     "3 to exit the program " +
                     "\n\nMake your entry here: ")

    if tryAgain == "1":
        theHash._updatedInfo_()

    if tryAgain == "2":
        theHash._getInfo_()

    if tryAgain == "3":
        print("\n----Exiting the Employee Secure Info Page----\n")
        exit()

    else:
        main()

main()

